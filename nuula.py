# -*- coding: utf-8 -*-

import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
import glob
import unittest
import os.path


def xmlparser(filedir):

  filepath = "Main.xlsx"
  wb = openpyxl.Workbook()

  wb.save(filepath)

  for file in glob.glob(filedir+"*.xml"):

    tree = ET.parse(file)
    root = tree.getroot()

    attrbs_error = [] 
    attrbs_messages = []
    attrbs_rules = []

    nuula_error = root.findall("./Nuula/Errors/Error")
    messages = root.findall("./DataExtract900jer/Messages/Message")
    rules = root.findall("./DataExtract900jer/Rules/Rule")
  
    if len(nuula_error) > 0:
      for err in nuula_error:
        attrbs_error.append(err.attrib)
      df_error = pd.DataFrame(attrbs_error)
    else:
      df_error = pd.DataFrame()

    if len(messages) > 0:
      for msg in messages:
        attrbs_messages.append(msg.attrib)
      df_messages = pd.DataFrame(attrbs_messages)
    else:
      df_messages = pd.DataFrame()

    if len(rules) > 0:
      for rl in rules:
        attrbs_rules.append(rl.attrib)
      df_rules = pd.DataFrame(attrbs_rules)
    else:
      df_rules = pd.DataFrame()

    with pd.ExcelWriter('Main.xlsx', engine='openpyxl', mode = 'a') as writer: 
      df_error.to_excel(writer, sheet_name=file[7:10]+"_Errors", index = False)
      df_messages.to_excel(writer, sheet_name=file[7:10]+"_Messages", index = False)
      df_rules.to_excel(writer, sheet_name=file[7:10]+"_Rules", index = False)
  
  wb = openpyxl.load_workbook('Main.xlsx')
  if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
  wb.save('Main.xlsx')


class MyTest(unittest.TestCase):

    def test_excel_creation(self):
        xmlparser("./data/")
        self.assertEqual(os.path.isfile('Main.xlsx'), True )

    def test_xml_exists(self):
      actual = False
      for fname in os.listdir('./data/'):
        if fname.endswith('.xml'):
          actual = True
          break
      self.assertEqual(actual, True )


if __name__ == '__main__':
    filedir = input("directory path : ")
    xmlparser(filedir)

